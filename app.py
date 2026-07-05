#!/usr/bin/env python3
import base64
import hashlib
import json
import mimetypes
import os
import shutil
import sqlite3
import threading
import uuid
from datetime import datetime
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
FRONTEND = ROOT / "frontend"
TEMPLATE_DIR = ROOT / "templates"
DATA_ROOT = Path(
    os.environ.get("APP_DATA_DIR")
    or os.environ.get("RAILWAY_VOLUME_MOUNT_PATH")
    or str(ROOT)
).resolve()
DB_PATH = DATA_ROOT / "db" / "qc_ir.sqlite3"
GENERATED_DIR = DATA_ROOT / "storage" / "generated"
UPLOAD_DIR = DATA_ROOT / "storage" / "uploads"

APP_TZ_LABEL = "Africa/Cairo"
SESSIONS = {}
DB_LOCK = threading.RLock()


PERMISSIONS = {
    "Admin": {
        "ADMIN_SETUP", "VIEW_DASHBOARD", "VIEW_QC", "SUBMIT_QC", "REVIEW_QC",
        "VIEW_IR", "UPDATE_IR_REPLY", "RESUBMIT_IR"
    },
    "Engineer": {"VIEW_DASHBOARD", "VIEW_QC", "SUBMIT_QC", "VIEW_IR"},
    "QC": {"VIEW_DASHBOARD", "VIEW_QC", "REVIEW_QC", "VIEW_IR", "RESUBMIT_IR"},
    "Document Controller": {"VIEW_DASHBOARD", "VIEW_QC", "VIEW_IR", "UPDATE_IR_REPLY"},
    "Manager": {"VIEW_DASHBOARD", "VIEW_QC", "VIEW_IR"},
}


def now_iso():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def today_iso():
    return datetime.now().strftime("%Y-%m-%d")


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def rows(cur):
    return [dict(r) for r in cur.fetchall()]


def password_hash(password):
    return hashlib.sha256(str(password).encode("utf-8")).hexdigest()


def normalize_flat(flat):
    value = str(flat or "").strip()
    return value.zfill(2) if value.isdigit() and len(value) == 1 else value


def duplicate_key(payload, item_no, stage_id):
    return "|".join([
        str(payload.get("building", "")).strip().upper(),
        str(payload.get("floor", "")).strip().upper(),
        normalize_flat(payload.get("flat", "")).upper(),
        str(payload.get("part", "")).strip().upper(),
        str(item_no).strip().upper(),
        str(stage_id).strip().upper(),
    ])


def active_qc_status(status):
    return status not in {"QC Rejected - Closed"}


def build_revision_serial(base, rev):
    return base if int(rev or 0) == 0 else f"{base}-{int(rev):02d}"


def next_serial(conn, key, pad):
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    current = int(row["value"] if row else 1)
    conn.execute(
        "INSERT INTO settings(key, value) VALUES(?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (key, str(current + 1)),
    )
    return str(current).zfill(pad)


def get_setting(conn, key, default=""):
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else default


def add_history(conn, linked_type, linked_id, action, old_status, new_status, comment, user_name):
    conn.execute(
        """
        INSERT INTO audit_history
        (history_id, linked_type, linked_id, action, old_status, new_status, comment, done_by, done_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (f"HIS-{uuid.uuid4().hex[:10]}", linked_type, linked_id, action, old_status, new_status, comment, user_name, now_iso()),
    )


def merged_parent(ws, cell_ref):
    for rng in ws.merged_cells.ranges:
        if cell_ref in rng:
            return rng.start_cell.coordinate
    return cell_ref


def set_cell(ws, cell_ref, value):
    ws[merged_parent(ws, cell_ref)] = value


def save_workbook(wb, prefix, serial):
    safe = "".join(ch if ch.isalnum() or ch in "-_." else "_" for ch in serial)
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    out = GENERATED_DIR / f"{prefix}_{safe}.xlsx"
    wb.save(out)
    return out


def public_file_key(path):
    resolved = Path(path).resolve()
    try:
        return str(resolved.relative_to(DATA_ROOT)).replace("\\", "/")
    except ValueError:
        return str(resolved.relative_to(ROOT)).replace("\\", "/")


def generate_qc_form(conn, qc):
    wb = load_workbook(TEMPLATE_DIR / "qc_template.xlsx")
    ws = wb.active

    code = str(qc.get("qc_code") or "").upper()
    check = "☑"
    clear = "☐"

    set_cell(ws, "D8", get_setting(conn, "CONTRACTOR_NAME", "Jidaar Construction And Development"))
    set_cell(ws, "M8", qc["qc_serial"])
    set_cell(ws, "D9", get_setting(conn, "PROJECT_NAME", "HAP (SLR)"))
    set_cell(ws, "M9", f"Rev{int(qc['qc_rev']):02d}")
    set_cell(ws, "D10", qc["discipline_name"])
    set_cell(ws, "M10", qc["submittal_date"])

    set_cell(ws, "B13", qc["subject"])
    set_cell(ws, "H13", qc["div"])
    set_cell(ws, "J13", qc["section"])
    set_cell(ws, "M13", qc["item_no"])
    set_cell(ws, "D19", f"{qc['building']} / {qc['floor']} / {qc['part']}{qc['flat']}")
    set_cell(ws, "K19", qc["stage_name"])
    set_cell(ws, "D21", qc["created_at"])
    set_cell(ws, "D23", f"Eng. {qc['created_by']}")
    set_cell(ws, "D25", f"Eng. {qc['created_by']}")
    set_cell(ws, "D27", qc["submittal_date"])

    set_cell(ws, "D32", check if code == "A" else clear)
    set_cell(ws, "G32", check if code == "B" else clear)
    set_cell(ws, "I32", check if code == "C" else clear)
    set_cell(ws, "K32", check if code == "D" else clear)
    if qc.get("qc_remark"):
        set_cell(ws, "B34", qc["qc_remark"])

    return save_workbook(wb, "QC", qc["qc_serial"])


def generate_ir_form(conn, qc, ir):
    wb = load_workbook(TEMPLATE_DIR / "ir_template.xlsx")
    ws = wb.active

    discipline = str(qc["discipline_code"] or "").upper()
    discipline_cells = {
        "CIVIL": "A6",
        "SURVEY": "C6",
        "ARCH": "E6",
        "ELECTRICAL": "H6",
        "MECHANICAL": "J6",
        "HVAC": "M6",
        "OTHER": "O6",
    }
    if discipline in discipline_cells:
        set_cell(ws, discipline_cells[discipline], f"☑ {ws[discipline_cells[discipline]].value}")

    set_cell(ws, "C8", get_setting(conn, "PARCEL", "6"))
    set_cell(ws, "G8", qc["location_text"])
    set_cell(ws, "M8", qc["part"])
    set_cell(ws, "F10", f"IR No:      {ir['serial_no']}")
    set_cell(ws, "M10", ir["submittal_date"])
    set_cell(ws, "M11", get_setting(conn, "DEFAULT_INSPECTION_TIME", ""))
    set_cell(ws, "K12", f"Inspector: {get_setting(conn, 'CONSULTANT_INSPECTOR', '')}")
    set_cell(ws, "C13", qc["floor"])
    set_cell(ws, "K13", f"Member: {qc['stage_name']}")
    set_cell(ws, "A14", f"Ref. BOQ: {qc['item_no']} - {qc['description']}")
    set_cell(ws, "A16", qc["div"])
    set_cell(ws, "F16", qc["section"])
    set_cell(ws, "L16", qc["item_no"])
    set_cell(ws, "A17", f"Inspection: {qc['subject']}")
    set_cell(ws, "A18", f"Attachments: {qc.get('attachment_note') or 'As attached in system'}")
    set_cell(ws, "C19", f"Eng/{qc['created_by']}")
    set_cell(ws, "C20", f"Eng/{get_setting(conn, 'DEFAULT_QC_ENGINEER', 'QC Engineer')}")
    set_cell(ws, "L19", ir["submittal_date"])
    set_cell(ws, "L20", ir["submittal_date"])

    return save_workbook(wb, "IR", ir["serial_no"])


def require_user(headers, permission=None):
    auth = headers.get("Authorization", "")
    token = auth.replace("Bearer ", "").strip()
    user = SESSIONS.get(token)
    if not user:
        raise ApiError(401, "Session expired. Please login again.")
    if permission and permission not in PERMISSIONS.get(user["role"], set()):
        raise ApiError(403, "Access denied for this action.")
    return user


class ApiError(Exception):
    def __init__(self, status, message):
        self.status = status
        self.message = message
        super().__init__(message)


def init_db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

    with DB_LOCK, db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
              user_id TEXT PRIMARY KEY,
              name TEXT NOT NULL,
              username TEXT UNIQUE NOT NULL,
              password_hash TEXT NOT NULL,
              role TEXT NOT NULL,
              discipline_code TEXT,
              is_active INTEGER NOT NULL DEFAULT 1,
              created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS settings (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS disciplines (
              discipline_code TEXT PRIMARY KEY,
              discipline_name TEXT NOT NULL,
              is_active INTEGER NOT NULL DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS locations (
              location_code TEXT PRIMARY KEY,
              location_name TEXT NOT NULL,
              is_active INTEGER NOT NULL DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS parts (
              part_code TEXT PRIMARY KEY,
              part_name TEXT NOT NULL,
              is_active INTEGER NOT NULL DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS boq_items (
              item_id TEXT PRIMARY KEY,
              buildings TEXT NOT NULL,
              item_no TEXT NOT NULL,
              description TEXT NOT NULL,
              discipline_code TEXT NOT NULL,
              discipline_name TEXT NOT NULL,
              div TEXT,
              section TEXT,
              unit TEXT,
              quantity REAL,
              subject_template TEXT NOT NULL,
              is_active INTEGER NOT NULL DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS item_stages (
              stage_id TEXT PRIMARY KEY,
              item_id TEXT NOT NULL,
              stage_no INTEGER NOT NULL,
              stage_name TEXT NOT NULL,
              sequence_required INTEGER NOT NULL DEFAULT 1,
              needs_ir INTEGER NOT NULL DEFAULT 1,
              is_active INTEGER NOT NULL DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS qc_requests (
              qc_id TEXT PRIMARY KEY,
              qc_base_serial TEXT NOT NULL,
              qc_rev INTEGER NOT NULL,
              qc_serial TEXT UNIQUE NOT NULL,
              subject TEXT NOT NULL,
              building TEXT NOT NULL,
              floor TEXT NOT NULL,
              flat TEXT NOT NULL,
              part TEXT NOT NULL,
              location_text TEXT NOT NULL,
              item_id TEXT NOT NULL,
              item_no TEXT NOT NULL,
              description TEXT NOT NULL,
              stage_id TEXT NOT NULL,
              stage_name TEXT NOT NULL,
              discipline_code TEXT NOT NULL,
              discipline_name TEXT NOT NULL,
              div TEXT,
              section TEXT,
              submittal_date TEXT NOT NULL,
              reply_date TEXT,
              qc_status TEXT NOT NULL,
              qc_code TEXT,
              qc_remark TEXT,
              current_status TEXT NOT NULL,
              created_by TEXT NOT NULL,
              reviewed_by TEXT,
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL,
              duplicate_key TEXT NOT NULL,
              qc_form_path TEXT,
              qc_email_status TEXT NOT NULL DEFAULT 'Pending Email',
              qc_reply_email_status TEXT NOT NULL DEFAULT 'Pending Email'
            );

            CREATE TABLE IF NOT EXISTS ir_requests (
              ir_id TEXT PRIMARY KEY,
              linked_qc_id TEXT NOT NULL,
              ir_base_serial TEXT NOT NULL,
              ir_rev INTEGER NOT NULL,
              serial_no TEXT UNIQUE NOT NULL,
              subject TEXT NOT NULL,
              location TEXT NOT NULL,
              building TEXT NOT NULL,
              floor TEXT NOT NULL,
              flat TEXT NOT NULL,
              part TEXT NOT NULL,
              discipline_code TEXT NOT NULL,
              discipline_name TEXT NOT NULL,
              div TEXT,
              section TEXT,
              item_no TEXT NOT NULL,
              stage_name TEXT NOT NULL,
              submittal_date TEXT NOT NULL,
              reply_date TEXT,
              code TEXT,
              remark TEXT,
              ir_status TEXT NOT NULL,
              created_by TEXT NOT NULL,
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL,
              duplicate_key TEXT NOT NULL,
              ir_form_path TEXT,
              ir_email_status TEXT NOT NULL DEFAULT 'Pending Email',
              ir_reply_email_status TEXT NOT NULL DEFAULT 'Pending Email'
            );

            CREATE TABLE IF NOT EXISTS attachments (
              attachment_id TEXT PRIMARY KEY,
              linked_type TEXT NOT NULL,
              linked_id TEXT NOT NULL,
              file_name TEXT NOT NULL,
              file_path TEXT NOT NULL,
              uploaded_by TEXT NOT NULL,
              uploaded_at TEXT NOT NULL,
              notes TEXT
            );

            CREATE TABLE IF NOT EXISTS audit_history (
              history_id TEXT PRIMARY KEY,
              linked_type TEXT NOT NULL,
              linked_id TEXT NOT NULL,
              action TEXT NOT NULL,
              old_status TEXT,
              new_status TEXT,
              comment TEXT,
              done_by TEXT NOT NULL,
              done_at TEXT NOT NULL
            );
            """
        )

        defaults = {
            "PROJECT_CODE": "SLR",
            "PACKAGE_CODE": "JID",
            "COMPANY_CODE": "ACE",
            "PHASE_CODE": "P06",
            "PARCEL": "6",
            "PROJECT_NAME": "HAP (SLR)",
            "CONTRACTOR_NAME": "Jidaar Construction And Development",
            "NEXT_QC_SERIAL": "1",
            "NEXT_IR_SERIAL": "9639",
            "DEFAULT_QC_ENGINEER": "Samah Atta",
            "DEFAULT_INSPECTION_TIME": "",
            "CONSULTANT_INSPECTOR": "",
        }
        for key, value in defaults.items():
            conn.execute("INSERT OR IGNORE INTO settings(key, value) VALUES (?, ?)", (key, value))

        seed_rows(conn)


def seed_rows(conn):
    conn.execute(
        "INSERT OR IGNORE INTO users VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        ("USR-admin", "System Admin", "admin", password_hash("admin123"), "Admin", "ARCH", 1, now_iso()),
    )
    for name, username, role, disc in [
        ("Site Architect", "arch", "Engineer", "ARCH"),
        ("QC Engineer", "qc", "QC", "ARCH"),
        ("Document Controller", "dc", "Document Controller", ""),
        ("Project Manager", "manager", "Manager", ""),
    ]:
        conn.execute(
            "INSERT OR IGNORE INTO users VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (f"USR-{username}", name, username, password_hash("123456"), role, disc, 1, now_iso()),
        )

    for code, name in [("ARCH", "Architect"), ("CIVIL", "Civil"), ("MECH", "Mechanical"), ("ELEC", "Electrical")]:
        conn.execute("INSERT OR IGNORE INTO disciplines VALUES (?, ?, 1)", (code, name))

    for code, name in [("B09", "Building 09"), ("B10", "Building 10"), ("BS", "Basement"), ("GF", "Ground Floor"), ("F01", "First Floor")]:
        conn.execute("INSERT OR IGNORE INTO locations VALUES (?, ?, 1)", (code, name))

    for code, name in [("APP", "Apartment"), ("CR", "Corridor & Lobbies"), ("ST", "Stairs"), ("SHF", "Shafts")]:
        conn.execute("INSERT OR IGNORE INTO parts VALUES (?, ?, 1)", (code, name))

    conn.execute(
        """
        INSERT OR IGNORE INTO boq_items
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
        """,
        (
            "ITM-ARCH-001", "B09,B10", "09 24 00",
            "Gypsum Board Works", "ARCH", "Architect", "09", "092400", "m2", 1000,
            "{Stage} for {Description} ({Building}-{Floor}-{Part}{Flat})",
        ),
    )

    for idx, stage in enumerate(["Material Inspection", "Installation Inspection", "Final Inspection"], start=1):
        conn.execute(
            "INSERT OR IGNORE INTO item_stages VALUES (?, ?, ?, ?, 1, 1, 1)",
            (f"STG-ARCH-001-{idx}", "ITM-ARCH-001", idx, stage),
        )


def build_subject(item, stage, payload):
    template = item["subject_template"]
    flat = normalize_flat(payload.get("flat"))
    return (
        template
        .replace("{Stage}", stage["stage_name"])
        .replace("{Description}", item["description"])
        .replace("{Building}", payload.get("building", ""))
        .replace("{Floor}", payload.get("floor", ""))
        .replace("{Part}", payload.get("part", ""))
        .replace("{Flat}", flat)
    )


def handle_login(body):
    with db() as conn:
        user = conn.execute(
            "SELECT * FROM users WHERE lower(username) = lower(?) AND password_hash = ? AND is_active = 1",
            (body.get("username"), password_hash(body.get("password", ""))),
        ).fetchone()
    if not user:
        raise ApiError(401, "Invalid username or password.")
    token = uuid.uuid4().hex
    user_dict = dict(user)
    user_dict.pop("password_hash", None)
    SESSIONS[token] = user_dict
    return {"ok": True, "token": token, "user": user_dict, "permissions": sorted(PERMISSIONS.get(user_dict["role"], set()))}


def handle_bootstrap(user):
    with db() as conn:
        return {
            "ok": True,
            "user": user,
            "permissions": sorted(PERMISSIONS.get(user["role"], set())),
            "settings": {r["key"]: r["value"] for r in conn.execute("SELECT * FROM settings").fetchall()},
            "disciplines": rows(conn.execute("SELECT * FROM disciplines WHERE is_active = 1 ORDER BY discipline_name")),
            "locations": rows(conn.execute("SELECT * FROM locations WHERE is_active = 1 ORDER BY location_code")),
            "parts": rows(conn.execute("SELECT * FROM parts WHERE is_active = 1 ORDER BY part_code")),
            "boq": rows(conn.execute("SELECT * FROM boq_items WHERE is_active = 1 ORDER BY item_no")),
            "stages": rows(conn.execute("SELECT * FROM item_stages WHERE is_active = 1 ORDER BY item_id, stage_no")),
            "qcLog": rows(conn.execute("SELECT * FROM qc_requests ORDER BY created_at DESC")),
            "irLog": rows(conn.execute("SELECT * FROM ir_requests ORDER BY created_at DESC")),
            "history": rows(conn.execute("SELECT * FROM audit_history ORDER BY done_at DESC LIMIT 100")),
        }


def submit_qc(user, body):
    with DB_LOCK, db() as conn:
        conn.execute("BEGIN IMMEDIATE")
        item = conn.execute("SELECT * FROM boq_items WHERE item_id = ? AND is_active = 1", (body.get("item_id"),)).fetchone()
        if not item:
            raise ApiError(400, "BOQ item not found.")
        if user["role"] == "Engineer" and item["discipline_code"].upper() != str(user.get("discipline_code") or "").upper():
            raise ApiError(403, "Engineer can submit only his discipline.")

        stage = conn.execute("SELECT * FROM item_stages WHERE stage_id = ? AND is_active = 1", (body.get("stage_id"),)).fetchone()
        if not stage:
            raise ApiError(400, "Stage not found.")

        payload = {
            "building": body.get("building", "").strip(),
            "floor": body.get("floor", "").strip(),
            "flat": normalize_flat(body.get("flat", "")),
            "part": body.get("part", "").strip(),
        }
        missing = [k for k, v in payload.items() if not v]
        if missing:
            raise ApiError(400, "Missing fields: " + ", ".join(missing))

        key = duplicate_key(payload, item["item_no"], stage["stage_id"])
        existing = conn.execute(
            "SELECT * FROM qc_requests WHERE duplicate_key = ? ORDER BY qc_rev DESC LIMIT 1",
            (key,),
        ).fetchone()
        if existing and active_qc_status(existing["current_status"]):
            raise ApiError(409, f"Already submitted by {existing['created_by']} as {existing['qc_serial']} ({existing['current_status']}).")

        if existing:
            base_serial = existing["qc_base_serial"]
            rev = int(existing["qc_rev"]) + 1
            serial = build_revision_serial(base_serial, rev)
            action = "SUBMIT_QC_REVISION"
        else:
            base_serial = "QC-" + next_serial(conn, "NEXT_QC_SERIAL", 4)
            rev = 0
            serial = base_serial
            action = "SUBMIT_QC"

        subject = build_subject(item, stage, payload)
        qc_id = f"QC-{uuid.uuid4().hex[:10]}"
        created_at = now_iso()
        conn.execute(
            """
            INSERT INTO qc_requests
            (qc_id, qc_base_serial, qc_rev, qc_serial, subject, building, floor, flat, part, location_text,
             item_id, item_no, description, stage_id, stage_name, discipline_code, discipline_name, div, section,
             submittal_date, reply_date, qc_status, qc_code, qc_remark, current_status, created_by, reviewed_by,
             created_at, updated_at, duplicate_key, qc_form_path)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '', ?, '', ?, ?, ?, '',
                    ?, ?, ?, ?)
            """,
            (
                qc_id, base_serial, rev, serial, subject, payload["building"], payload["floor"], payload["flat"], payload["part"],
                f"{payload['building']} / {payload['floor']} / {payload['part']}{payload['flat']}",
                item["item_id"], item["item_no"], item["description"], stage["stage_id"], stage["stage_name"],
                item["discipline_code"], item["discipline_name"], item["div"], item["section"], today_iso(),
                "Pending QC", body.get("notes", ""), "Pending QC", user["name"], created_at, created_at, key, "",
            ),
        )
        qc = dict(conn.execute("SELECT * FROM qc_requests WHERE qc_id = ?", (qc_id,)).fetchone())
        form_path = generate_qc_form(conn, qc)
        conn.execute("UPDATE qc_requests SET qc_form_path = ? WHERE qc_id = ?", (public_file_key(form_path), qc_id))
        add_history(conn, "QC", qc_id, action, "", "Pending QC", body.get("notes", ""), user["name"])
        conn.commit()
        return {"ok": True, "message": f"QC submitted: {serial}", "qc": dict(conn.execute("SELECT * FROM qc_requests WHERE qc_id = ?", (qc_id,)).fetchone())}


def approve_qc(user, qc_id, body):
    code = str(body.get("code", "A")).strip().upper()
    if code not in {"A", "B"}:
        raise ApiError(400, "Approval code must be A or B.")

    with DB_LOCK, db() as conn:
        conn.execute("BEGIN IMMEDIATE")
        qc = conn.execute("SELECT * FROM qc_requests WHERE qc_id = ?", (qc_id,)).fetchone()
        if not qc:
            raise ApiError(404, "QC not found.")
        if qc["current_status"] not in {"Pending QC"}:
            raise ApiError(409, f"QC is already {qc['current_status']}.")

        status = "Approved" if code == "A" else "Approved with Comments"
        conn.execute(
            """
            UPDATE qc_requests
            SET reply_date = ?, qc_status = ?, qc_code = ?, qc_remark = ?, current_status = 'QC Approved',
                reviewed_by = ?, updated_at = ?, qc_reply_email_status = 'Pending Email'
            WHERE qc_id = ?
            """,
            (today_iso(), status, code, body.get("remark", ""), user["name"], now_iso(), qc_id),
        )
        qc = dict(conn.execute("SELECT * FROM qc_requests WHERE qc_id = ?", (qc_id,)).fetchone())
        qc_form = generate_qc_form(conn, qc)
        conn.execute("UPDATE qc_requests SET qc_form_path = ? WHERE qc_id = ?", (public_file_key(qc_form), qc_id))
        add_history(conn, "QC", qc_id, "APPROVE_QC", "Pending QC", "QC Approved", f"Code {code}", user["name"])

        existing_ir = conn.execute("SELECT * FROM ir_requests WHERE linked_qc_id = ?", (qc_id,)).fetchone()
        if existing_ir:
            ir = dict(existing_ir)
        else:
            settings = {r["key"]: r["value"] for r in conn.execute("SELECT * FROM settings").fetchall()}
            base = "-".join([
                settings.get("PROJECT_CODE", "SLR"),
                settings.get("PACKAGE_CODE", "JID"),
                settings.get("COMPANY_CODE", "ACE"),
                "IR",
                settings.get("PHASE_CODE", "P06"),
                qc["discipline_code"],
                next_serial(conn, "NEXT_IR_SERIAL", 4),
            ])
            ir_id = f"IR-{uuid.uuid4().hex[:10]}"
            conn.execute(
                """
                INSERT INTO ir_requests
                (ir_id, linked_qc_id, ir_base_serial, ir_rev, serial_no, subject, location, building, floor, flat, part,
                 discipline_code, discipline_name, div, section, item_no, stage_name, submittal_date, reply_date, code,
                 remark, ir_status, created_by, created_at, updated_at, duplicate_key, ir_form_path)
                VALUES (?, ?, ?, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '', '', '', 'IR Submitted',
                        ?, ?, ?, ?, '')
                """,
                (
                    ir_id, qc_id, base, base, qc["subject"], qc["location_text"], qc["building"], qc["floor"], qc["flat"], qc["part"],
                    qc["discipline_code"], qc["discipline_name"], qc["div"], qc["section"], qc["item_no"], qc["stage_name"],
                    today_iso(), user["name"], now_iso(), now_iso(), qc["duplicate_key"],
                ),
            )
            ir = dict(conn.execute("SELECT * FROM ir_requests WHERE ir_id = ?", (ir_id,)).fetchone())
            ir_form = generate_ir_form(conn, qc, ir)
            conn.execute("UPDATE ir_requests SET ir_form_path = ? WHERE ir_id = ?", (public_file_key(ir_form), ir_id))
            conn.execute("UPDATE qc_requests SET current_status = 'IR Generated', updated_at = ? WHERE qc_id = ?", (now_iso(), qc_id))
            add_history(conn, "IR", ir_id, "GENERATE_IR", "", "IR Submitted", base, user["name"])
            ir = dict(conn.execute("SELECT * FROM ir_requests WHERE ir_id = ?", (ir_id,)).fetchone())

        conn.commit()
        return {"ok": True, "message": f"QC approved and IR generated: {ir['serial_no']}", "ir": ir}


def reject_qc(user, qc_id, body):
    code = str(body.get("code", "C")).strip().upper()
    remark = str(body.get("remark", "")).strip()
    if not remark:
        raise ApiError(400, "Reject remark is required.")

    with DB_LOCK, db() as conn:
        conn.execute("BEGIN IMMEDIATE")
        qc = conn.execute("SELECT * FROM qc_requests WHERE qc_id = ?", (qc_id,)).fetchone()
        if not qc:
            raise ApiError(404, "QC not found.")
        if qc["current_status"] != "Pending QC":
            raise ApiError(409, f"QC is already {qc['current_status']}.")
        conn.execute(
            """
            UPDATE qc_requests
            SET reply_date = ?, qc_status = 'Rejected', qc_code = ?, qc_remark = ?,
                current_status = 'QC Rejected - Closed', reviewed_by = ?, updated_at = ?,
                qc_reply_email_status = 'Pending Email'
            WHERE qc_id = ?
            """,
            (today_iso(), code, remark, user["name"], now_iso(), qc_id),
        )
        qc = dict(conn.execute("SELECT * FROM qc_requests WHERE qc_id = ?", (qc_id,)).fetchone())
        form_path = generate_qc_form(conn, qc)
        conn.execute("UPDATE qc_requests SET qc_form_path = ? WHERE qc_id = ?", (public_file_key(form_path), qc_id))
        add_history(conn, "QC", qc_id, "REJECT_QC", "Pending QC", "QC Rejected - Closed", f"Code {code}: {remark}", user["name"])
        conn.commit()
        return {"ok": True, "message": "QC rejected. Next submit will create a revision."}


def update_ir_reply(user, ir_id, body):
    code = str(body.get("code", "")).strip().upper()
    remark = str(body.get("remark", "")).strip()
    if code not in {"A", "B", "C", "D"}:
        raise ApiError(400, "IR reply code must be A, B, C, or D.")
    status = "IR Approved" if code in {"A", "B"} else "IR Rejected - Closed"
    with DB_LOCK, db() as conn:
        conn.execute("BEGIN IMMEDIATE")
        ir = conn.execute("SELECT * FROM ir_requests WHERE ir_id = ?", (ir_id,)).fetchone()
        if not ir:
            raise ApiError(404, "IR not found.")
        old = ir["ir_status"]
        conn.execute(
            """
            UPDATE ir_requests
            SET reply_date = ?, code = ?, remark = ?, ir_status = ?, updated_at = ?,
                ir_reply_email_status = 'Pending Email'
            WHERE ir_id = ?
            """,
            (today_iso(), code, remark, status, now_iso(), ir_id),
        )
        add_history(conn, "IR", ir_id, "UPDATE_IR_REPLY", old, status, f"Code {code}: {remark}", user["name"])
        conn.commit()
        return {"ok": True, "message": f"IR updated: {status}"}


def handle_sync(user, body):
    results = []
    for op in body.get("operations", []):
        try:
            kind = op.get("type")
            payload = op.get("payload", {})
            if kind == "SUBMIT_QC":
                result = submit_qc(user, payload)
            elif kind == "APPROVE_QC":
                result = approve_qc(user, payload.get("qc_id"), payload)
            elif kind == "REJECT_QC":
                result = reject_qc(user, payload.get("qc_id"), payload)
            else:
                raise ApiError(400, f"Unknown offline operation: {kind}")
            results.append({"client_id": op.get("client_id"), "ok": True, "result": result})
        except ApiError as err:
            results.append({"client_id": op.get("client_id"), "ok": False, "status": err.status, "message": err.message})
        except Exception as err:
            results.append({"client_id": op.get("client_id"), "ok": False, "status": 500, "message": str(err)})
    return {"ok": True, "results": results}


class Handler(SimpleHTTPRequestHandler):
    def translate_path(self, path):
        parsed = urlparse(path)
        if parsed.path.startswith("/files/"):
            requested = parsed.path.replace("/files/", "", 1).lstrip("/")
            target = (DATA_ROOT / requested).resolve()
            allowed_roots = [
                GENERATED_DIR.resolve(),
                UPLOAD_DIR.resolve(),
            ]
            if not any(str(target).startswith(str(base)) for base in allowed_roots):
                return str(ROOT / "__not_allowed__")
            return str(target)
        target = parsed.path.lstrip("/") or "index.html"
        frontend_target = (FRONTEND / target).resolve()
        if not str(frontend_target).startswith(str(FRONTEND.resolve())):
            return str(FRONTEND / "index.html")
        return str(frontend_target)

    def end_headers(self):
        self.send_header("Cache-Control", "no-store" if self.path.startswith("/api/") else "public, max-age=60")
        super().end_headers()

    def send_json(self, data, status=200):
        raw = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def read_json(self):
        length = int(self.headers.get("Content-Length", "0"))
        if length == 0:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/health":
            return self.send_json({"ok": True, "time": now_iso(), "timezone": APP_TZ_LABEL})
        if parsed.path == "/api/bootstrap":
            try:
                return self.send_json(handle_bootstrap(require_user(self.headers)))
            except ApiError as err:
                return self.send_json({"ok": False, "message": err.message}, err.status)
        if parsed.path.startswith("/files/"):
            return super().do_GET()
        if not (FRONTEND / parsed.path.lstrip("/")).exists() and "." not in parsed.path:
            self.path = "/index.html"
        return super().do_GET()

    def do_POST(self):
        parsed = urlparse(self.path)
        try:
            body = self.read_json()
            if parsed.path == "/api/login":
                return self.send_json(handle_login(body))
            if parsed.path == "/api/qc":
                return self.send_json(submit_qc(require_user(self.headers, "SUBMIT_QC"), body))
            if parsed.path.endswith("/approve") and parsed.path.startswith("/api/qc/"):
                qc_id = parsed.path.split("/")[3]
                return self.send_json(approve_qc(require_user(self.headers, "REVIEW_QC"), qc_id, body))
            if parsed.path.endswith("/reject") and parsed.path.startswith("/api/qc/"):
                qc_id = parsed.path.split("/")[3]
                return self.send_json(reject_qc(require_user(self.headers, "REVIEW_QC"), qc_id, body))
            if parsed.path.endswith("/reply") and parsed.path.startswith("/api/ir/"):
                ir_id = parsed.path.split("/")[3]
                return self.send_json(update_ir_reply(require_user(self.headers, "UPDATE_IR_REPLY"), ir_id, body))
            if parsed.path == "/api/offline/sync":
                return self.send_json(handle_sync(require_user(self.headers), body))
            raise ApiError(404, "Endpoint not found.")
        except ApiError as err:
            return self.send_json({"ok": False, "message": err.message}, err.status)
        except Exception as err:
            return self.send_json({"ok": False, "message": str(err)}, 500)


def main():
    init_db()
    port = int(os.environ.get("PORT", "8080"))
    server = ThreadingHTTPServer(("0.0.0.0", port), Handler)
    print(f"QC/IR SaaS running at http://localhost:{port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
